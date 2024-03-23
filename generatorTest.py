import openpyxl
import locale
from datetime import datetime

def get_data(sheet, clinic_name, row):
    adspend = sheet.cell(row=row, column=2).value
    cpl = sheet.cell(row=row, column=3).value
    leads = sheet.cell(row=row, column=4).value
    return adspend, cpl, leads

def update_data(sheet, clinic_name, row, dates):
    # Set the appropriate locale for reading prices with commas
    locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8' if ',' in str(sheet.cell(row=row, column=2).value) else 'en_US')

    if clinic_name is not None and clinic_name != 'Datum':
        while True:
            try:
                adspend = float(input(f"Enter Ad Spend for {clinic_name}: €").replace(',', ''))
                cpl = float(input(f"Enter CPL for {clinic_name}: €").replace(',', ''))
                break
            except ValueError:
                print("Invalid input. Please enter a valid number.")

        # Reset the locale to avoid affecting other parts of the program
        locale.setlocale(locale.LC_NUMERIC, 'C')

        # Calculate the number of leads
        leads = int(adspend / cpl)

        sheet.cell(row=row, column=2).value = adspend
        sheet.cell(row=row, column=3).value = cpl
        sheet.cell(row=row, column=4).value = leads

        # Format cells to display Euro sign (excluding the leads column)
        for col in range(2, 4):
            sheet.cell(row=row, column=col).number_format = '€#,##0.00'

        # Store the date in the dates list
        dates[row - 2] = sheet.cell(row=row, column=9).value

def get_date_input(prompt):
    while True:
        try:
            date_str = input(prompt + " (YYYY-MM-DD): ")
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj
        except ValueError:
            print("Invalid date format. Please enter a date in YYYY-MM-DD format.")

def main():
    # Load the Excel file
    wb = openpyxl.load_workbook('cosmetics.xlsx')
    sheet = wb.active

    # Get user input for date range
    start_date = get_date_input("Enter start date")
    end_date = get_date_input("Enter end date")

    # Write start date and end date to cells B9 and C9
    sheet.cell(row=9, column=2).value = start_date.strftime("%Y-%m-%d")
    sheet.cell(row=9, column=3).value = end_date.strftime("%Y-%m-%d")

    # Assuming the clinic names are in column A and the dates in column 9
    clinic_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    dates = [sheet.cell(row=i, column=9).value for i in range(2, sheet.max_row + 1)]

    for i, clinic_name in enumerate(clinic_names, start=2):
        update_data(sheet, clinic_name, i, dates)

    # Save the updated workbook
    wb.save('cosmetics_updated.xlsx')
    print("Data successfully updated and saved to cosmetics_updated.xlsx.")

    # Explicitly close the workbook
    wb.close()

if __name__ == "__main__":
    main()