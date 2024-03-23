import openpyxl
import locale

def get_data(sheet, clinic_name, row):
    adspend = sheet.cell(row=row, column=2).value
    cpl = sheet.cell(row=row, column=3).value
    leads = sheet.cell(row=row, column=4).value
    return adspend, cpl, leads

def update_data(sheet, clinic_name, row, dates):
    # Set the appropriate locale for reading prices with commas
    locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8' if ',' in str(sheet.cell(row=row, column=2).value) else 'en_US')

    if clinic_name is not None and clinic_name != 'Datum':
        while True:
            try:
                adspend = float(input(f"Enter Ad Spend for {clinic_name}: €").replace(',', ''))
                cpl = float(input(f"Enter CPL for {clinic_name}: €").replace(',', ''))
                break
            except ValueError:
                print("Invalid input. Please enter a valid number.")

        # Reset the locale to avoid affecting other parts of the program
        locale.setlocale(locale.LC_NUMERIC, 'C')

        # Calculate the number of leads
        leads = int(adspend / cpl)

        sheet.cell(row=row, column=2).value = adspend
        sheet.cell(row=row, column=3).value = cpl
        sheet.cell(row=row, column=4).value = leads

        # Format cells to display Euro sign (excluding the leads column)
        for col in range(2, 4):
            sheet.cell(row=row, column=col).number_format = '€#,##0.00'

        # Store the date in the dates list
        dates[row - 2] = sheet.cell(row=row, column=9).value

def main():
    # Load the Excel file
    wb = openpyxl.load_workbook('cosmetics.xlsx')
    sheet = wb.active

    # Assuming the clinic names are in column A and the dates in column 9
    clinic_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    dates = [sheet.cell(row=i, column=9).value for i in range(2, sheet.max_row + 1)]

    for i, clinic_name in enumerate(clinic_names, start=2):
        update_data(sheet, clinic_name, i, dates)

    # Set the locale to ensure correct date parsing
    # locale.s

    # Save the updated workbook
    wb.save('cosmetics_updated.xlsx')
    print("Data successfully updated and saved to cosmetics_updated.xlsx.")

if __name__ == "__main__":
    main()
